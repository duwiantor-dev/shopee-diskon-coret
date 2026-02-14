import io
import re
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell


# =========================
# TEMPLATE (sesuai request baru)
# =========================
MASS_HEADER_ROW_FIXED = 1
MASS_DATA_START_ROW_FIXED = 2

PRICELIST_HEADER_ROW_FIXED = 2  # tetap

# =========================
# Konstanta header (Mass Update)
# =========================
MASS_HEADER_SKU = "SKU Ref. No.(Optional)"
MASS_HEADER_PRICE = "Harga diskon"

# Pricelist: header minimal yang kita cari
PL_HEADER_SKU_CANDIDATES = ["KODEBARANG", "KODE BARANG", "SKU", "SKU NO", "SKU_NO", "KODEBARANG "]
PL_PRICE_COL_TIKTOK = "M3"
PL_PRICE_COL_SHOPEE = "M4"

# Addon mapping: header yang diharapkan (boleh lebih dari 1 kandidat)
ADDON_CODE_CANDIDATES = ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"]
ADDON_PRICE_CANDIDATES = ["harga", "HARGA", "Price", "PRICE", "Harga"]

# Heuristik: jika nilai < 1.000.000 dianggap "tanpa 000" dan perlu dikali 1000
SMALL_TO_THOUSAND_THRESHOLD = 1_000_000
AUTO_MULTIPLIER_FOR_SMALL = 1000

SKU_SPLIT_PLUS = re.compile(r"\+")


# =========================
# Utils
# =========================
def normalize_text(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def normalize_addon_code(x) -> str:
    return normalize_text(x).upper()


def detect_platform_from_filename(filename: str) -> str:
    # Placeholder (harga selalu pakai M4)
    return "shopee"


def parse_platform_sku(full_sku: str) -> Tuple[str, List[str]]:
    if full_sku is None:
        return "", []

    s = str(full_sku).strip()
    if not s:
        return "", []

    parts = SKU_SPLIT_PLUS.split(s)
    base = parts[0].strip()
    addons = [p.strip() for p in parts[1:] if p and p.strip()]
    return base, addons


def parse_number_like_id(x) -> str:
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if x.is_integer():
            return str(int(x))
        return str(x)
    return str(x).strip()


def parse_price_cell(val) -> Optional[int]:
    if val is None:
        return None

    if isinstance(val, (int, float)):
        try:
            if isinstance(val, float) and val.is_integer():
                return int(val)
            return int(round(float(val)))
        except Exception:
            return None

    s = str(val).strip()
    if not s:
        return None

    s = s.replace("Rp", "").replace("rp", "").replace(" ", "")

    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s and "," not in s:
        s = s.replace(".", "")
    elif "," in s and "." not in s:
        s = s.replace(",", "")

    try:
        f = float(s)
        if f.is_integer():
            return int(f)
        return int(round(f))
    except Exception:
        return None


def apply_multiplier_if_needed(x: int) -> int:
    if x is None:
        return 0
    if x < SMALL_TO_THOUSAND_THRESHOLD:
        return x * AUTO_MULTIPLIER_FOR_SMALL
    return x


def safe_set_cell_value(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        coord = cell.coordinate
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        return
    cell.value = value


# =========================
# Excel scanning (FIXED ROW)
# =========================
def find_header_row_and_cols_mass(ws) -> Tuple[int, int, int]:
    """
    FIX: header Mass Update ada di row 1.
    Return: (header_row_idx, sku_col_idx, price_col_idx)
    """
    r = MASS_HEADER_ROW_FIXED
    target_a = MASS_HEADER_SKU.strip().lower()
    target_b = MASS_HEADER_PRICE.strip().lower()

    row_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        row_vals.append("" if v is None else str(v).strip())

    lower_to_col = {}
    for idx, v in enumerate(row_vals, start=1):
        lv = v.strip().lower()
        if lv and lv not in lower_to_col:
            lower_to_col[lv] = idx

    if target_a not in lower_to_col or target_b not in lower_to_col:
        raise ValueError(
            f"Header Mass Update row {MASS_HEADER_ROW_FIXED} tidak sesuai. "
            f"Pastikan ada '{MASS_HEADER_SKU}' dan '{MASS_HEADER_PRICE}'."
        )

    return r, lower_to_col[target_a], lower_to_col[target_b]


def find_header_row_and_cols_pricelist(ws) -> Tuple[int, int, int, int]:
    """
    FIX: header Pricelist ada di row 2.
    Return: (header_row_idx, sku_col_idx, m3_col_idx, m4_col_idx)
    """
    r = PRICELIST_HEADER_ROW_FIXED
    candidates = [c.strip().lower() for c in PL_HEADER_SKU_CANDIDATES]
    target_m3 = PL_PRICE_COL_TIKTOK.lower()
    target_m4 = PL_PRICE_COL_SHOPEE.lower()

    row_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        row_vals.append("" if v is None else str(v).strip())

    lower_to_col = {}
    for idx, v in enumerate(row_vals, start=1):
        lv = v.strip().lower()
        if lv and lv not in lower_to_col:
            lower_to_col[lv] = idx

    sku_col = None
    for cand in candidates:
        if cand in lower_to_col:
            sku_col = lower_to_col[cand]
            break

    if sku_col is None or target_m3 not in lower_to_col or target_m4 not in lower_to_col:
        raise ValueError(
            f"Header Pricelist row {PRICELIST_HEADER_ROW_FIXED} tidak sesuai. "
            f"Pastikan ada kolom KODEBARANG (atau setara) dan kolom M3 & M4."
        )

    return r, sku_col, lower_to_col[target_m3], lower_to_col[target_m4]


def load_pricelist_map(pl_bytes: bytes) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    ws = wb.active

    header_row, sku_col, m3_col, m4_col = find_header_row_and_cols_pricelist(ws)

    m: Dict[str, Dict[str, int]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku_val = ws.cell(row=r, column=sku_col).value
        sku = normalize_text(sku_val)
        if not sku:
            continue

        m3_raw = parse_price_cell(ws.cell(row=r, column=m3_col).value)
        m4_raw = parse_price_cell(ws.cell(row=r, column=m4_col).value)

        if m3_raw is None and m4_raw is None:
            continue

        m3 = apply_multiplier_if_needed(m3_raw) if m3_raw is not None else None
        m4 = apply_multiplier_if_needed(m4_raw) if m4_raw is not None else None

        m[sku] = {}
        if m3 is not None:
            m[sku]["M3"] = int(m3)
        if m4 is not None:
            m[sku]["M4"] = int(m4)

    return m


def load_addon_map(addon_bytes: bytes) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active

    header_row = None
    code_col = None
    price_col = None

    code_cands = [c.strip().lower() for c in ADDON_CODE_CANDIDATES]
    price_cands = [c.strip().lower() for c in ADDON_PRICE_CANDIDATES]

    for r in range(1, 30):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v).strip())

        lower_to_col = {}
        for idx, v in enumerate(row_vals, start=1):
            lv = v.strip().lower()
            if not lv:
                continue
            if lv not in lower_to_col:
                lower_to_col[lv] = idx

        found_code = None
        for cc in code_cands:
            if cc in lower_to_col:
                found_code = lower_to_col[cc]
                break

        found_price = None
        for pc in price_cands:
            if pc in lower_to_col:
                found_price = lower_to_col[pc]
                break

        if found_code and found_price:
            header_row = r
            code_col = found_code
            price_col = found_price
            break

    if header_row is None or code_col is None or price_col is None:
        raise ValueError("Header Addon Mapping tidak ketemu. Pastikan ada kolom addon_code & harga (atau setara).")

    m: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = normalize_addon_code(ws.cell(row=r, column=code_col).value)
        if not code:
            continue

        price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
        if price_raw is None:
            continue

        price = apply_multiplier_if_needed(int(price_raw))
        m[code] = int(price)

    return m


@dataclass
class RowChange:
    file: str
    excel_row: int
    sku_full: str
    old_price: int
    new_price: int
    reason: str


def compute_new_price_for_row(
    sku_full: str,
    platform: str,  # placeholder
    pl_map: Dict[str, Dict[str, int]],
    addon_map: Dict[str, int],
    discount_rp: int,
) -> Tuple[Optional[int], str]:
    base_sku, addons = parse_platform_sku(sku_full)
    if not base_sku:
        return None, "SKU kosong"

    pl = pl_map.get(base_sku)
    if not pl:
        return None, "Base SKU tidak ada di Pricelist"

    price_key = "M4"  # ✅ SELALU M4
    base_price = pl.get(price_key)
    if base_price is None:
        return None, f"Harga {price_key} kosong di Pricelist"

    addon_total = 0
    for a in addons:
        code = normalize_addon_code(a)
        if not code:
            continue
        if code not in addon_map:
            return None, f"Addon '{code}' tidak ada di file Addon Mapping"
        addon_total += int(addon_map[code])

    final_price = int(base_price) + int(addon_total) - int(discount_rp)
    if final_price < 0:
        final_price = 0

    return final_price, f"{price_key} + addon - diskon"


def make_issues_workbook(changes: List[RowChange]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "issues_report"
    headers = ["file", "row", "sku_full", "old_price", "new_price", "reason"]
    ws.append(headers)
    for ch in changes:
        ws.append([ch.file, ch.excel_row, ch.sku_full, ch.old_price, ch.new_price, ch.reason])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def workbook_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def keep_only_changed_rows_in_place(ws, data_start_row: int, changed_row_numbers: List[int]):
    """
    Hapus baris yang TIDAK berubah mulai dari data_start_row ke bawah,
    supaya format template tetap aman.
    """
    keep = set(changed_row_numbers)
    for r in range(ws.max_row, data_start_row - 1, -1):
        if r not in keep:
            ws.delete_rows(r, 1)


# =========================
# UI
# =========================
st.set_page_config(page_title="Discount Nominate (Shopee M4)", layout="wide")
st.title("Discount Nominate (Shopee M4)")

c1, c2, c3 = st.columns(3)
with c1:
    mass_files = st.file_uploader(
        "Upload Template Mass Update (bisa banyak)",
        type=["xlsx"],
        accept_multiple_files=True,
    )
with c2:
    pl_file = st.file_uploader("Upload Pricelist", type=["xlsx"])
with c3:
    addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"])

st.divider()

discount_rp = st.number_input("Diskon (Rp) - mengurangi harga final", min_value=0, value=0, step=1000)
process = st.button("Proses")

if process:
    if not mass_files or pl_file is None or addon_file is None:
        st.error("Wajib upload: Template Mass Update (minimal 1), Pricelist, dan Addon Mapping.")
        st.stop()

    try:
        pl_map = load_pricelist_map(pl_file.getvalue())
    except Exception as e:
        st.error(f"Gagal baca Pricelist: {e}")
        st.stop()

    try:
        addon_map = load_addon_map(addon_file.getvalue())
    except Exception as e:
        st.error(f"Gagal baca Addon Mapping: {e}")
        st.stop()

    changed_rows: List[RowChange] = []
    output_files: List[Tuple[str, bytes]] = []

    for mf in mass_files:
        filename = mf.name
        platform = detect_platform_from_filename(filename)

        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active

        try:
            header_row, sku_col, price_col = find_header_row_and_cols_mass(ws)
        except Exception as e:
            changed_rows.append(RowChange(
                file=filename,
                excel_row=0,
                sku_full="",
                old_price=0,
                new_price=0,
                reason=f"Gagal baca header mass update: {e}",
            ))
            continue

        changed_row_numbers: List[int] = []

        for r in range(MASS_DATA_START_ROW_FIXED, ws.max_row + 1):
            sku_val = ws.cell(row=r, column=sku_col).value
            sku_full = parse_number_like_id(sku_val)
            if not sku_full:
                continue

            old_price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
            old_price = int(old_price_raw) if old_price_raw is not None else 0

            new_price, reason = compute_new_price_for_row(
                sku_full=sku_full,
                platform=platform,
                pl_map=pl_map,
                addon_map=addon_map,
                discount_rp=int(discount_rp),
            )

            if new_price is None:
                continue
            if int(new_price) == int(old_price):
                continue

            safe_set_cell_value(ws, row=r, col=price_col, value=int(new_price))
            changed_row_numbers.append(r)

            changed_rows.append(RowChange(
                file=filename,
                excel_row=r,
                sku_full=sku_full,
                old_price=int(old_price),
                new_price=int(new_price),
                reason=reason,
            ))

        if changed_row_numbers:
            keep_only_changed_rows_in_place(
                ws,
                data_start_row=MASS_DATA_START_ROW_FIXED,
                changed_row_numbers=changed_row_numbers,
            )
            out_bytes = workbook_to_bytes(wb)
            out_name = filename.replace(".xlsx", "_changed_only_M4.xlsx")
            output_files.append((out_name, out_bytes))

    st.subheader("Preview (yang berubah saja)")
    if not changed_rows:
        st.info("Tidak ada perubahan harga.")
    else:
        import pandas as pd
        df_preview = pd.DataFrame([{
            "file": x.file,
            "row": x.excel_row,
            "sku_full": x.sku_full,
            "old_price": x.old_price,
            "new_price": x.new_price,
            "reason": x.reason,
        } for x in changed_rows])
        st.dataframe(df_preview.head(200), use_container_width=True)

    st.divider()

    if len(output_files) == 0:
        st.warning("Tidak ada file yang berubah, jadi tidak ada file output untuk didownload.")
    elif len(output_files) == 1:
        name, data = output_files[0]
        st.download_button(
            "Download hasil (XLSX) - hanya yang berubah (format asli)",
            data=data,
            file_name=name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        zbuf = io.BytesIO()
        with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, data in output_files:
                zf.writestr(name, data)
            rep = make_issues_workbook(changed_rows)
            zf.writestr("changes_report.xlsx", rep)

        st.download_button(
            "Download semua hasil (ZIP) - hanya yang berubah (format asli)",
            data=zbuf.getvalue(),
            file_name="mass_update_results_changed_only_M4.zip",
            mime="application/zip",
        )

    if changed_rows:
        rep_bytes = make_issues_workbook(changed_rows)
        st.download_button(
            "Download Report Perubahan (XLSX)",
            data=rep_bytes,
            file_name="changes_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


